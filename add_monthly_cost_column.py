import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE_PATH = r'C:\Users\User\OneDrive\Desktop\SeekWander_Infrastructure_Ledger.xlsx'

# Load workbook preserving all existing formatting
wb = openpyxl.load_workbook(FILE_PATH)
ws = wb.active

# ── Helper style factories ──────────────────────────────────────────────────

def header_font():
    return Font(name='Arial', size=10, bold=True, color='FFFFFFFF')

def header_fill():
    return PatternFill(fill_type='solid', fgColor='FFC2410C')

def header_alignment():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def thin_border(color='FFD0D0D0'):
    side = Side(border_style='thin', color=color)
    return Border(left=side, right=side, top=side, bottom=side)

def data_font():
    return Font(name='Arial', size=10)

def data_fill_white():
    return PatternFill(fill_type='solid', fgColor='FFFFFFFF')

def data_fill_paper():
    return PatternFill(fill_type='solid', fgColor='FFF5F0E8')

def data_alignment():
    return Alignment(horizontal='left', vertical='top', wrap_text=True)

# ── Cost data for rows 5–22 ─────────────────────────────────────────────────

cost_data = {
    5:  "Free tier: Hobby (free, 100GB bandwidth). Pro: $20/seat/month. Recommended for production.",
    6:  "Free tier: 500MB DB, 2GB bandwidth. Pro: $25/month (8GB DB, 50GB bandwidth). Paused after 1 week inactivity on free tier.",
    7:  "Free — open-source ORM. No API key or subscription required.",
    8:  "Free tier: up to 10,000 MAUs. Pro: $25/month (unlimited MAUs + advanced features).",
    9:  "Pay-per-token. claude-sonnet-4-6: ~$3.00/M input tokens · ~$15.00/M output tokens. No monthly minimum.",
    10: "$200/month free credit (shared across all Google Maps APIs). Map loads: $7.00/1,000 after credit.",
    11: "$17.00/1,000 requests (after $200 free monthly credit, shared across all Google APIs).",
    12: "$32.00/1,000 requests (after $200 free monthly credit).",
    13: "$17.00/1,000 requests (after $200 free monthly credit).",
    14: "$17.00/1,000 requests (after $200 free monthly credit).",
    15: "$7.00/1,000 requests (after $200 free monthly credit).",
    16: "$10.00/1,000 elements (after $200 free monthly credit). Each origin-destination pair = 1 element.",
    17: "Free — served via Google's CDN. No usage limits or API key required.",
    18: "Free — open-source MIT licensed npm package. No subscription required.",
    19: "Free — open-source ISC licensed npm package. No subscription required.",
    20: "Free — open-source MIT licensed build-time tool. No subscription required.",
    21: "Free — open-source MIT licensed framework by Vercel. No own subscription.",
    22: "Free tier: 50 requests/hour (Demo). Production: $0 with attribution, or Unsplash+ for commercial use at custom pricing. Currently not in active use.",
}

# ── Step 1: Extend title and subtitle merges from A:D to A:E ────────────────

# Unmerge existing A1:D1 and A2:D2, then re-merge to A1:E1 and A2:E2
for merge_range in ['A1:D1', 'A2:D2']:
    if merge_range in [str(m) for m in ws.merged_cells.ranges]:
        ws.unmerge_cells(merge_range)

ws.merge_cells('A1:E1')
ws.merge_cells('A2:E2')

# ── Step 2: Add header cell E4 ──────────────────────────────────────────────

e4 = ws.cell(row=4, column=5)
e4.value = "MONTHLY COST / PRICING"
e4.font = header_font()
e4.fill = header_fill()
e4.alignment = header_alignment()
e4.border = thin_border()

# ── Step 3: Set column E width ───────────────────────────────────────────────

ws.column_dimensions['E'].width = 38

# ── Step 4: Add cost data to rows 5–22 ──────────────────────────────────────

for row_num, text in cost_data.items():
    cell = ws.cell(row=row_num, column=5)
    cell.value = text
    cell.font = data_font()
    # Alternating fill: odd rows (5,7,9...) = white, even rows (6,8,10...) = paper
    if row_num % 2 == 1:
        cell.fill = data_fill_white()
    else:
        cell.fill = data_fill_paper()
    cell.alignment = data_alignment()
    cell.border = thin_border()

# ── Step 5: Extend footer merge from A23:D23 to A23:E23 ─────────────────────

if 'A23:D23' in [str(m) for m in ws.merged_cells.ranges]:
    ws.unmerge_cells('A23:D23')

ws.merge_cells('A23:E23')

# The footer cell (A23) already has the correct content and styling — no changes needed.

# ── Step 6: Save ─────────────────────────────────────────────────────────────

wb.save(FILE_PATH)
print(f"Saved: {FILE_PATH}")

# ── Step 7: Verify ───────────────────────────────────────────────────────────

assert os.path.exists(FILE_PATH), "ERROR: File not found after save!"
file_size = os.path.getsize(FILE_PATH)
print(f"Verification OK — file exists, size: {file_size:,} bytes")

# Quick sanity checks
wb2 = openpyxl.load_workbook(FILE_PATH)
ws2 = wb2.active
assert ws2.cell(row=4, column=5).value == "MONTHLY COST / PRICING", "Header E4 mismatch"
assert ws2.cell(row=5, column=5).value.startswith("Free tier: Hobby"), "Row 5 cost mismatch"
assert ws2.cell(row=9, column=5).value.startswith("Pay-per-token"), "Row 9 cost mismatch"
assert ws2.cell(row=22, column=5).value.startswith("Free tier: 50 requests"), "Row 22 cost mismatch"
assert 'A1:E1' in [str(m) for m in ws2.merged_cells.ranges], "Title merge A1:E1 missing"
assert 'A2:E2' in [str(m) for m in ws2.merged_cells.ranges], "Subtitle merge A2:E2 missing"
assert 'A23:E23' in [str(m) for m in ws2.merged_cells.ranges], "Footer merge A23:E23 missing"
print("All sanity checks passed.")
print(f"Merged cells: {[str(m) for m in ws2.merged_cells.ranges]}")
